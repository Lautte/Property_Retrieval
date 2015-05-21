__author__ = 'arthurl'
import re #https://docs.python.org/2/howto/regex.html
#import shutil
#import os
import xlrd
from xlrd import open_workbook
from openpyxl import load_workbook
import openpyxl
import datetime

### Open 'Read' workbook with xlrd.  NOTE:  XLRD can only read from xls files!!!###
workbook = xlrd.open_workbook('JOHN_appendix_B.xls')
### Read data from source column using xlrd ###

#Grab a specific worksheet from a workbook using xlrd
worksheet = workbook.sheet_by_name('Table 1')
#r=0
acreage=[]
survey=[]
ROE=[]
ADDRESS=[]
for rownum in range(worksheet.nrows):
    #try:
        row_val = str(worksheet.cell_value(rownum, 0))
        p = re.compile("ACREAGE: SURVEY:")
        m = p.search(row_val)
        if m:
            print 'match found: ', m.group()
            survey.append(str(worksheet.cell_value(rownum+1, 0)))
            acreage.append(str(worksheet.cell_value(rownum+2,0)))
            print "survey No. "+str(worksheet.cell_value(rownum+1, 0))
            #i=1

            #This portion of the code finds the right of entry data, and uses the
                #"TRACT NO:" as a marker to bump the program INTO this
                #parsing loop.  This portion needs to be reworked, because it is only
                #catching 68 out of 75 of the ROE
            """roe_row=(str(worksheet.cell_value(rownum+i,0)))
            ROE_comp=re.compile("TRACT NO:")
            ROE_search=ROE_comp.search(roe_row)
            while ROE_search:
                roe=[]
                for k in [-1,-2,-3,-4,-5,-6,-7,-8,-9,-10,-11,-12,-13,-14,-15,-16]:
                    notes_val=str(worksheet.cell_value(rownum+k+i,0))
                    notes=re.compile('^'+"RIGHT OF ENTRY:")
                    seek=notes.search(notes_val)
                    if seek:
                        print "Right O' Entry "+seek.group()
                        roe.append(str(worksheet.cell_value(rownum+k+i,0)))
                        print "---------"
                        break

                    else:
                            #print str(worksheet.cell_value(rownum+k+i,0))
                        roe.append(str(worksheet.cell_value(rownum+k+i,0)))
                rightofentry=roe[::-1]#this reverses the list here so it is in the correct order
                rightofentry=' '.join(rightofentry)

                print rightofentry
                ROE.append(rightofentry)
                print "---"
                break"""
            for i in [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40]:
                 #This portion of the code finds the right of entry data, and uses the
                #"TRACT NO:" as a marker to bump the program INTO this
                #parsing loop.  This portion needs to be reworked, because it is only
                #catching 68 out of 75 of the ROE
                roe_row=(str(worksheet.cell_value(rownum+i,0)))
                ROE_comp=re.compile("TRACT NO:")
                ROE_search=ROE_comp.search(roe_row)
                if  ROE_search:
                    roe=[]
                    for k in [-1,-2,-3,-4,-5,-6,-7,-8,-9,-10,-11,-12,-13,-14,-15,-16]:
                        notes_val=str(worksheet.cell_value(rownum+k+i,0))
                        notes=re.compile('^'+"RIGHT OF ENTRY:{1,2}")
                        seek=notes.search(notes_val)
                        if seek:
                            print "Right O' Entry "+seek.group()
                            roe.append(str(worksheet.cell_value(rownum+k+i,0)))
                            print "---------"
                            break

                        else:
                            #print str(worksheet.cell_value(rownum+k+i,0))
                            roe.append(str(worksheet.cell_value(rownum+k+i,0)))
                    rightofentry=roe[::-1]#this reverses the list here so it is in the correct order
                    rightofentry=' '.join(rightofentry)

                    print rightofentry
                    ROE.append(rightofentry)
                    print "---"
                    break
                    #ROE.append(str(worksheet.cell_value(rownum+i+1,0)))

                #This portion of the code finds the surface owners address, and uses the
                #"surface resident and address" as a marker to bump the program out of this
                #parsing loop
                elif str(worksheet.cell_value(rownum+i,0)) == "SURFACE OWNERS AND ADDRESS:":
                    address=[]
                    for j in [1,2,3,4,5,6,7,8,9,10]:
                        if str(worksheet.cell_value(rownum+j+i,0))== "SURFACE RESIDENT AND ADDRESS:":
                            break
                        else:
                            #print str(worksheet.cell_value(rownum+j+i,0))
                            #print "--"
                            address.append(str(worksheet.cell_value(rownum+j+i,0)))
                    addr=' '.join(address)
                    print addr
                    ADDRESS.append(addr)
                    #print "----"

                else:
                    continue
                    #print str(worksheet.cell_value(rownum+i,0))

        elif not m:
            print 'No Match for row: '+row_val
            continue
    #except(UnicodeEncodeError):
     #   print "Unicode Encode Error on line:  "+str(rownum+1)
        #ROE.append('No ROE, b/c of error')
      #  pass
#print acreage
#print survey
#print ROE
"""o_ls = zip(survey, acreage, ADDRESS)
def create_dict(data):
    from collections import defaultdict
    ret = defaultdict(list)
    for v in data:
        ret[v[0]].append(v[1:])
    return ret

output=dict(create_dict(o_ls))
#print output
#print o_ls
################################################ Open destination workbook using openpyxl AND WRITE TO IT USING OPENPYXL ###
destwb = load_workbook('JOHN_appendix_B_EDIT.xlsx')
#wksht = destwb.create_sheet(0)
#print str(wksht)
ws=destwb.get_sheet_by_name('Sheet1')
rw=1
print "LENGTH OF List of Tuples "+str(len(o_ls))
print "Length of Acreage List "+ str(len(acreage))
print "Length of Address List "+ str(len(ADDRESS))
print "Length of ROE List "+ str(len(ROE))
while rw < len(survey):
    for i, (sur, ac, add) in enumerate(o_ls):
        ws.cell(column=1, row=rw).value=sur
        ws.cell(column=2, row=rw).value=ac
        ws.cell(column=3, row=rw).value=add
        #ws.cell(column=4, row=rw).value=right
        #print sur
        #print ac
        rw+=1


destwb.save('PROPERTY_EDIT.xlsx')"""

