__author__ = 'arthurl'
import re #https://docs.python.org/2/howto/regex.html
#import shutil
#import os
import xlrd
from xlrd import open_workbook
from openpyxl import load_workbook
import openpyxl
import datetime

### Open 'Read' workbook with xlrd.  NOTE:  XLRD can only read from xls files!!!###
workbook = xlrd.open_workbook('Cleanedup_PDF.xls')
### Read data from source column using xlrd ###

#Grab a specific worksheet from a workbook using xlrd
worksheet = workbook.sheet_by_name('Table 1')
#r=0
ACREAGE=[]
TRACTNO=[]
ROE=[]
ADDRESS=[]
for rownum in range(worksheet.nrows):
    try:
        row_val = str(worksheet.cell_value(rownum, 0))

        q2 = re.compile("^"+"END"+"$")#I added an "END" on the last row of the Cleanedup_PDF.xls
        THEEND = q2.search(row_val)

        p = re.compile('^'+"TRACT NO:")
        m = p.search(row_val)
        if m:
            #print 'match found: ', m.group()
            TRACTNO.append(str(worksheet.cell_value(rownum, 0)))
            #print rownum

            ACREAGE.append(str(worksheet.cell_value(rownum+1,0)))

            print "BEGINNING "+str(worksheet.cell_value(rownum, 0))
            #i=1

            for i in [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40]:
                 #This portion of the code finds the right of entry data, and uses the
                #"TRACT NO:" as a marker to bump the program INTO this
                #parsing loop.
                #print i

                roe_row=(str(worksheet.cell_value(rownum+i,0)))
                #print roe_row
                the_end=re.compile('^'+"END"+'$')
                theend=the_end.search(roe_row)
                ROE_comp=re.compile('^'+"TRACT NO:")
                ROE_search=ROE_comp.search(roe_row)
                ADDRESS_comp=re.compile('^'+"SURFACE OWNERS AND ADDRESS:")
                ADDRESS_search=ADDRESS_comp.search(roe_row)
                if  ROE_search:
                    roe=[]
                    for k in [-1,-2,-3,-4,-5,-6,-7,-8,-9,-10,-11,-12,-13,-14,-15,-16,-17,-18,-19,-20,-21,-22,-23,-24,-25]:
                        notes_val=str(worksheet.cell_value(rownum+k+i,0))
                        notes=re.compile('^'+"RIGHT OF ENTRY:{1,2}")
                        seek=notes.search(notes_val)
                        if seek:
                            print "Right O' Entry "+seek.group()
                            roe.append(str(worksheet.cell_value(rownum+k+i,0)))
                            #print "---------"
                            break

                        else:
                            #print str(worksheet.cell_value(rownum+k+i,0))
                            roe.append(str(worksheet.cell_value(rownum+k+i,0)))
                    rightofentry=roe[::-1]#this reverses the list here so it is in the correct order
                    rightofentry=' '.join(rightofentry)

                    #print rightofentry
                    ROE.append(rightofentry)
                    #print "---"
                    #print "TRACT No. END "+str(worksheet.cell_value(rownum+i, 0))
                    break

                    #ROE.append(str(worksheet.cell_value(rownum+i+1,0)))

                #This portion of the code finds the surface owners address, and uses the
                #"surface resident and address" as a marker to bump the program out of this
                #parsing loop
                elif ADDRESS_search:
                    address=[]

                    for j in [1,2,3,4,5,6,7,8,9,10]:
                        addr_row=str(worksheet.cell_value(rownum+i+j,0))
                        ADDR_comp=re.compile('^'+"SURFACE RESIDENT AND ADDRESS:")
                        ADDR_search=ADDR_comp.search(addr_row)
                        if ADDR_search:
                            break
                        else:
                            #print str(worksheet.cell_value(rownum+j+i,0))
                            #print "--"
                            address.append(str(worksheet.cell_value(rownum+j+i,0)))
                    addr=' '.join(address)
                    print "ADDRESS IS: "+addr
                    ADDRESS.append(addr)
                    #print "----"
                elif theend:#this sees the 'END' statement and gathers the ROE data for the last Property Tract
                    print "THE END sorta"
                    roe=[]
                    for k1 in [-1,-2,-3,-4,-5,-6,-7,-8,-9,-10,-11,-12,-13,-14,-15,-16,-17,-18,-19,-20,-21,-22,-23,-24,-25]:
                        notes_val=str(worksheet.cell_value(rownum+k1+i,0))
                        notes=re.compile('^'+"RIGHT OF ENTRY:")
                        seek=notes.search(notes_val)
                        if seek:
                            print "Right O' Entry "+seek.group()
                            roe.append(str(worksheet.cell_value(rownum+k1+i,0)))
                            break
                        else:
                            #print str(worksheet.cell_value(rownum+k+i,0))
                            roe.append(str(worksheet.cell_value(rownum+k1+i,0)))
                    rightofentry=roe[::-1]#this reverses the list here so it is in the correct order
                    rightofentry=' '.join(rightofentry)
                    ROE.append(rightofentry)
                    break

                else:
                    continue
                    #print str(worksheet.cell_value(rownum+i,0))
            #print "TRACT No. END "+str(worksheet.cell_value(rownum+i, 0))

        elif not m:
            #print 'No Match for row: '+row_val
            continue
        #elif THEEND:
         #   print "THE END Fo rizzle"
          #  break
    except(UnicodeEncodeError):
        print "Unicode Encode Error on line:  "+str(rownum+1)
        ROE.append('No ROE, b/c of error')
        pass
print "Length of TractNo List "+ str(len(TRACTNO))
print "Length of Acreage List "+ str(len(ACREAGE))
print "Length of Address List "+ str(len(ADDRESS))
print "Length of ROE List "+ str(len(ROE))
o_ls = zip(TRACTNO, ACREAGE,ADDRESS, ROE)
################################################ Open destination workbook using openpyxl AND WRITE TO IT USING OPENPYXL ###
destwb = load_workbook('JOHN_appendix_B_EDIT.xlsx')
ws=destwb.get_sheet_by_name('Sheet1')
rw=1
#print "LENGTH OF List of Tuples "+str(len(o_ls))

while rw < len(TRACTNO):
    for i, (tract, ac,ADDr, RoE) in enumerate(o_ls):
        ws.cell(column=1, row=rw).value=tract
        ws.cell(column=2, row=rw).value=ac
        ws.cell(column=3, row=rw).value=ADDr
        ws.cell(column=4, row=rw).value=RoE
        #print sur
        #print ac
        rw+=1


destwb.save('PROPERTY_EDIT.xlsx')